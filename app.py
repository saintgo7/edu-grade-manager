import math
import io
import csv
import os
import re
from datetime import datetime

import xlrd

from flask import (
    Flask, render_template, request, redirect, url_for, flash, abort
)
from flask_sqlalchemy import SQLAlchemy

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///grade_management.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class Course(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    year = db.Column(db.Integer, nullable=False)
    semester = db.Column(db.String(20), nullable=False)
    report_weight = db.Column(db.Float, default=20.0)
    attendance_weight = db.Column(db.Float, default=10.0)
    midterm_weight = db.Column(db.Float, default=35.0)
    final_weight = db.Column(db.Float, default=35.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    students = db.relationship(
        'Student', backref='course', lazy=True, cascade='all, delete-orphan'
    )
    grade_policy = db.relationship(
        'GradePolicy', backref='course', uselist=False, cascade='all, delete-orphan'
    )

    def __repr__(self):
        return f'<Course {self.name}>'


class GradePolicy(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    a_plus  = db.Column(db.Float, default=10.0)
    a       = db.Column(db.Float, default=15.0)
    b_plus  = db.Column(db.Float, default=15.0)
    b       = db.Column(db.Float, default=20.0)
    c_plus  = db.Column(db.Float, default=15.0)
    c       = db.Column(db.Float, default=10.0)
    d_plus  = db.Column(db.Float, default=5.0)
    d       = db.Column(db.Float, default=5.0)
    # F covers the remainder automatically

    def __repr__(self):
        return f'<GradePolicy course_id={self.course_id}>'


class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    student_number = db.Column(db.String(50))   # 학번 (nullable — 명단 없이 추가 가능)
    department = db.Column(db.String(100))       # 계열 (nullable)
    report_score = db.Column(db.Float, nullable=False)
    attendance_score = db.Column(db.Float, nullable=False)
    midterm_score = db.Column(db.Float, nullable=False)
    final_score = db.Column(db.Float, nullable=False)
    total_score = db.Column(db.Float)
    grade = db.Column(db.String(2))
    rank = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Student {self.name}>'


class CourseRoster(db.Model):
    """명단 — 명단 Excel에서 가져온 학생 목록. Student와 별개로 관리."""
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), nullable=False)
    department = db.Column(db.String(100))   # 계열
    name = db.Column(db.String(100), nullable=False)  # 이름
    student_number = db.Column(db.String(50))  # 학번
    course = db.relationship(
        'Course',
        backref=db.backref('roster', lazy=True, cascade='all, delete-orphan'),
    )

    def __repr__(self):
        return f'<CourseRoster {self.name}>'


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def calc_total_score(student, course):
    """Return weighted total (0-100 scale)."""
    return (
        student.report_score * course.report_weight
        + student.attendance_score * course.attendance_weight
        + student.midterm_score * course.midterm_weight
        + student.final_score * course.final_weight
    ) / 100


def recalculate_grades(course_id):
    """Recalculate total_score, rank and grade for every student in the course."""
    course = db.session.get(Course, course_id)
    if course is None:
        return
    policy = course.grade_policy
    if policy is None:
        return

    # First update every student's total_score
    for student in course.students:
        student.total_score = calc_total_score(student, course)

    # Sort by total → midterm → final → report → attendance (all descending)
    students = Student.query.filter_by(course_id=course_id).order_by(
        Student.total_score.desc(),
        Student.midterm_score.desc(),
        Student.final_score.desc(),
        Student.report_score.desc(),
        Student.attendance_score.desc(),
    ).all()

    n = len(students)
    if n == 0:
        db.session.commit()
        return

    counts = [
        math.ceil(n * policy.a_plus  / 100),
        math.ceil(n * policy.a       / 100),
        math.ceil(n * policy.b_plus  / 100),
        math.ceil(n * policy.b       / 100),
        math.ceil(n * policy.c_plus  / 100),
        math.ceil(n * policy.c       / 100),
        math.ceil(n * policy.d_plus  / 100),
        math.ceil(n * policy.d       / 100),
    ]
    grade_labels = ['A+', 'A', 'B+', 'B', 'C+', 'C', 'D+', 'D']

    named_total = sum(counts)
    idx = 0
    for rank, student in enumerate(students, 1):
        student.rank = rank
        cumulative = sum(counts[:idx + 1])
        while idx < len(counts) - 1 and rank > cumulative:
            idx += 1
            cumulative = sum(counts[:idx + 1])
        if rank > named_total:
            student.grade = 'F'
        else:
            student.grade = grade_labels[idx]

    db.session.commit()


def _parse_score(raw, field_name):
    """Parse raw string to float in [0, 100]. Returns (value, error_msg)."""
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return None, f'{field_name} 점수는 숫자여야 합니다.'
    if not (0.0 <= v <= 100.0):
        return None, f'{field_name} 점수는 0~100 사이여야 합니다.'
    return v, None


def _validate_weights(report_w, attend_w, midterm_w, final_w):
    """Return error string if weights don't sum to 100, else None."""
    total = report_w + attend_w + midterm_w + final_w
    if abs(total - 100.0) > 0.01:
        return f'가중치 합계가 {total:.1f}%입니다. 합계가 정확히 100%여야 합니다.'
    return None


def _validate_policy(a_plus, a, b_plus, b, c_plus, c, d_plus, d):
    """Return error string if policy bands exceed 100, else None."""
    total = a_plus + a + b_plus + b + c_plus + c + d_plus + d
    if total > 100.0 + 0.01:
        return f'학점 구간 합계가 {total:.1f}%입니다. 100% 이하여야 합니다.'
    return None


def parse_roster_file(file_data, filename):
    """
    Parse a roster Excel file (.xls or .xlsx).

    Returns a list of dicts with keys: department, name, student_number.
    Raises ValueError with a user-facing message on any parsing error.

    Column detection is header-name based (position-independent):
      계열 / department   → department
      이름 / name         → name
      학번 / student_id / student_number → student_number
    """
    fname = filename.lower()
    if fname.endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            app.logger.exception('openpyxl parse error')
            raise ValueError('파일을 읽는 중 오류가 발생했습니다. 올바른 .xlsx 파일인지 확인해주세요.')
    elif fname.endswith('.xls'):
        try:
            wb = xlrd.open_workbook(file_contents=file_data)
            ws = wb.sheet_by_index(0)
            rows = [ws.row_values(r) for r in range(ws.nrows)]
        except Exception:
            app.logger.exception('xlrd parse error')
            raise ValueError('파일을 읽는 중 오류가 발생했습니다. 올바른 .xls 파일인지 확인해주세요.')
    else:
        raise ValueError('.xlsx 또는 .xls 파일만 업로드할 수 있습니다.')

    if not rows:
        raise ValueError('파일에 데이터가 없습니다.')

    def _cell_val(val):
        """Normalize a raw cell value to a clean string."""
        if val is None:
            return ''
        if isinstance(val, float):
            if not math.isfinite(val):
                return ''
            return str(int(val)) if val == int(val) else str(val)
        return str(val).strip()

    # ── 출석부 형식 감지 ──────────────────────────────────────────────
    # 출석부는 컬럼 헤더가 없고, 학생 데이터가 2행 쌍으로 반복됨:
    #   홀수 행: 계열 (예: "공학3계열")
    #   짝수 행: 이름 (학번) (예: "강세종 (20263153)")
    # A열(index 0) 에 "이름 (8자리 이상 숫자)" 패턴이 있으면 출석부로 판단.
    ATTENDANCE_PATTERN = re.compile(r'^(.+?)\s*\((\d{6,})\)\s*$')

    def _is_attendance_format(rows):
        for row in rows:
            if not row:
                continue
            cell = _cell_val(row[0])
            if ATTENDANCE_PATTERN.match(cell):
                return True
        return False

    if _is_attendance_format(rows):
        # 출석부 파싱: A열만 사용, 2행 쌍 (계열 / 이름(학번)) 반복
        results = []
        dept_buf = ''
        for row in rows:
            if not row:
                continue
            cell = _cell_val(row[0])
            if not cell:
                continue
            m = ATTENDANCE_PATTERN.match(cell)
            if m:
                # 이름(학번) 행
                name = m.group(1).strip()
                student_number = m.group(2).strip()
                results.append({
                    'department': dept_buf,
                    'name': name,
                    'student_number': student_number,
                })
                dept_buf = ''
            else:
                # 계열 행 (또는 헤더) — 다음 이름 행의 계열로 버퍼링
                # 헤더 잡음("학과\n이름(학번)", "2026학년도..." 등)은 길이/패턴으로 필터
                if len(cell) <= 30 and not any(k in cell for k in ('학년도', '출석부', '교과목', '담당', '수업', '학수', '서명', '주차', '분반')):
                    dept_buf = cell
        if not results:
            raise ValueError('출석부에서 학생 데이터를 찾을 수 없습니다. 형식을 확인해주세요.')
        return results

    # ── 표준 형식: 헤더 컬럼 방식 ────────────────────────────────────
    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    dept_idx = name_idx = num_idx = None
    for i, h in enumerate(header):
        if h in ('계열', 'department'):
            dept_idx = i
        elif h in ('이름', 'name'):
            name_idx = i
        elif h in ('학번', 'student_id', 'student_number'):
            num_idx = i

    if name_idx is None:
        raise ValueError(
            '헤더에서 이름 컬럼을 찾을 수 없습니다. '
            '"이름" 또는 "name" 컬럼이 있는지 확인해주세요. '
            '출석부 파일은 "이름 (학번)" 형식(예: 홍길동 (20230001))으로 되어 있어야 합니다.'
        )

    results = []
    for row in rows[1:]:
        if all((c is None or str(c).strip() == '') for c in row):
            continue

        def _cell(idx):
            if idx is None or idx >= len(row):
                return ''
            return _cell_val(row[idx])

        name = _cell(name_idx)
        if not name:
            continue
        results.append({
            'department': _cell(dept_idx),
            'name': name,
            'student_number': _cell(num_idx),
        })

    return results


def _parse_course_form(form):
    """
    Parse and validate course + grade policy from a form POST.
    Returns (data_dict, error_message). On error data_dict is None.
    """
    name = form.get('name', '').strip()
    if not name:
        return None, '과목명을 입력해주세요.'

    try:
        year = int(form.get('year', 0))
    except ValueError:
        return None, '연도는 정수여야 합니다.'

    semester = form.get('semester', '').strip()
    if semester not in ('1학기', '2학기', '여름', '겨울'):
        return None, '올바른 학기를 선택해주세요.'

    # weights
    w_fields = [
        ('report_weight',     '레포트'),
        ('attendance_weight', '출석'),
        ('midterm_weight',    '중간시험'),
        ('final_weight',      '기말시험'),
    ]
    weights = {}
    for field, label in w_fields:
        val, err = _parse_score(form.get(field, ''), label + ' 가중치')
        if err:
            return None, err
        weights[field] = val

    err = _validate_weights(
        weights['report_weight'], weights['attendance_weight'],
        weights['midterm_weight'], weights['final_weight']
    )
    if err:
        return None, err

    # grade policy bands
    p_fields = [
        ('a_plus',  'A+'),
        ('a',       'A'),
        ('b_plus',  'B+'),
        ('b',       'B'),
        ('c_plus',  'C+'),
        ('c',       'C'),
        ('d_plus',  'D+'),
        ('d',       'D'),
    ]
    policy = {}
    for field, label in p_fields:
        val, err = _parse_score(form.get(field, ''), label + ' 구간')
        if err:
            return None, err
        policy[field] = val

    err = _validate_policy(**policy)
    if err:
        return None, err

    return {
        'name': name,
        'year': year,
        'semester': semester,
        **weights,
        **policy,
    }, None


# ---------------------------------------------------------------------------
# Routes — Course
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    """Course list."""
    courses = Course.query.order_by(Course.year.desc(), Course.created_at.desc()).all()
    return render_template('index.html', courses=courses)


@app.route('/course/new', methods=['GET', 'POST'])
def course_new():
    """Create a new course with weights and grade policy."""
    if request.method == 'POST':
        data, err = _parse_course_form(request.form)
        if err:
            flash(err, 'error')
            return render_template('course_new.html', form=request.form)

        course = Course(
            name=data['name'],
            year=data['year'],
            semester=data['semester'],
            report_weight=data['report_weight'],
            attendance_weight=data['attendance_weight'],
            midterm_weight=data['midterm_weight'],
            final_weight=data['final_weight'],
        )
        db.session.add(course)
        db.session.flush()  # get course.id before committing

        policy = GradePolicy(
            course_id=course.id,
            a_plus=data['a_plus'],
            a=data['a'],
            b_plus=data['b_plus'],
            b=data['b'],
            c_plus=data['c_plus'],
            c=data['c'],
            d_plus=data['d_plus'],
            d=data['d'],
        )
        db.session.add(policy)
        db.session.commit()

        flash(f'과목 "{course.name}"이 생성되었습니다.', 'success')
        return redirect(url_for('course_detail', course_id=course.id))

    return render_template('course_new.html', form={})


@app.route('/course/<int:course_id>')
def course_detail(course_id):
    """Student list with ranks for a course."""
    course = Course.query.get_or_404(course_id)
    students = Student.query.filter_by(course_id=course_id).order_by(
        Student.rank.asc()
    ).all()
    return render_template('course_detail.html', course=course, students=students)


@app.route('/course/<int:course_id>/edit', methods=['GET', 'POST'])
def course_edit(course_id):
    """Edit course name, weights, and grade policy."""
    course = Course.query.get_or_404(course_id)
    policy = course.grade_policy

    if request.method == 'POST':
        data, err = _parse_course_form(request.form)
        if err:
            flash(err, 'error')
            return render_template('course_edit.html', course=course, policy=policy, form=request.form)

        course.name = data['name']
        course.year = data['year']
        course.semester = data['semester']
        course.report_weight = data['report_weight']
        course.attendance_weight = data['attendance_weight']
        course.midterm_weight = data['midterm_weight']
        course.final_weight = data['final_weight']

        policy.a_plus  = data['a_plus']
        policy.a       = data['a']
        policy.b_plus  = data['b_plus']
        policy.b       = data['b']
        policy.c_plus  = data['c_plus']
        policy.c       = data['c']
        policy.d_plus  = data['d_plus']
        policy.d       = data['d']

        db.session.flush()
        recalculate_grades(course.id)

        flash(f'과목 "{course.name}" 설정이 저장되었습니다.', 'success')
        return redirect(url_for('course_detail', course_id=course.id))

    return render_template('course_edit.html', course=course, policy=policy, form={})


@app.route('/course/<int:course_id>/delete', methods=['POST'])
def course_delete(course_id):
    """Delete a course and all its students (cascade)."""
    course = Course.query.get_or_404(course_id)
    name = course.name
    db.session.delete(course)
    db.session.commit()
    flash(f'과목 "{name}"이 삭제되었습니다.', 'success')
    return redirect(url_for('index'))


# ---------------------------------------------------------------------------
# Routes — Student
# ---------------------------------------------------------------------------

@app.route('/course/<int:course_id>/add_student', methods=['GET', 'POST'])
def add_student(course_id):
    """Add a student with 4 scores to a course."""
    course = Course.query.get_or_404(course_id)

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('학생 이름을 입력해주세요.', 'error')
            return render_template('add_student.html', course=course, form=request.form)

        score_fields = [
            ('report_score',     '레포트'),
            ('attendance_score', '출석'),
            ('midterm_score',    '중간시험'),
            ('final_score',      '기말시험'),
        ]
        scores = {}
        for field, label in score_fields:
            val, err = _parse_score(request.form.get(field, ''), label)
            if err:
                flash(err, 'error')
                return render_template('add_student.html', course=course, form=request.form)
            scores[field] = val

        student_number = request.form.get('student_number', '').strip() or None
        department = request.form.get('department', '').strip() or None

        student = Student(
            course_id=course_id,
            name=name,
            student_number=student_number,
            department=department,
            **scores,
        )
        db.session.add(student)
        db.session.commit()
        recalculate_grades(course_id)

        flash(f'{name} 학생이 추가되었습니다.', 'success')
        return redirect(url_for('course_detail', course_id=course_id))

    return render_template('add_student.html', course=course, form={})


@app.route('/student/<int:student_id>/edit', methods=['GET', 'POST'])
def edit_student(student_id):
    """Edit a student's 4 scores."""
    student = Student.query.get_or_404(student_id)
    course = student.course

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('학생 이름을 입력해주세요.', 'error')
            return render_template('edit_student.html', student=student, course=course, form=request.form)

        score_fields = [
            ('report_score',     '레포트'),
            ('attendance_score', '출석'),
            ('midterm_score',    '중간시험'),
            ('final_score',      '기말시험'),
        ]
        scores = {}
        for field, label in score_fields:
            val, err = _parse_score(request.form.get(field, ''), label)
            if err:
                flash(err, 'error')
                return render_template('edit_student.html', student=student, course=course, form=request.form)
            scores[field] = val

        student.name = name
        student.student_number = request.form.get('student_number', '').strip() or None
        student.department = request.form.get('department', '').strip() or None
        student.report_score = scores['report_score']
        student.attendance_score = scores['attendance_score']
        student.midterm_score = scores['midterm_score']
        student.final_score = scores['final_score']

        db.session.commit()
        recalculate_grades(course.id)

        flash(f'{name} 학생 정보가 수정되었습니다.', 'success')
        return redirect(url_for('course_detail', course_id=course.id))

    return render_template('edit_student.html', student=student, course=course, form={})


@app.route('/student/<int:student_id>/delete', methods=['POST'])
def delete_student(student_id):
    """Delete a student and recalculate grades."""
    student = Student.query.get_or_404(student_id)
    course_id = student.course_id
    name = student.name
    db.session.delete(student)
    db.session.commit()
    recalculate_grades(course_id)
    flash(f'{name} 학생이 삭제되었습니다.', 'success')
    return redirect(url_for('course_detail', course_id=course_id))


# ---------------------------------------------------------------------------
# Routes — Bulk Add (CSV)
# ---------------------------------------------------------------------------

@app.route('/course/<int:course_id>/bulk_add', methods=['GET', 'POST'])
def bulk_add(course_id):
    """Bulk-add students via CSV text (name,report,attendance,midterm,final)."""
    course = Course.query.get_or_404(course_id)

    if request.method == 'POST':
        raw = request.form.get('students_data', '').strip()
        if not raw:
            flash('학생 데이터를 입력해주세요.', 'error')
            return render_template('bulk_add.html', course=course)

        lines = raw.splitlines()
        added = 0
        errors = []

        for i, line in enumerate(lines, 1):
            line = line.strip()
            if not line:
                continue
            # Skip header if present
            if i == 1 and (line.lower().startswith('이름') or line.lower().startswith('name')):
                continue

            parts = [p.strip() for p in line.split(',')]
            if len(parts) != 5:
                errors.append(f'라인 {i}: 열이 5개여야 합니다 (이름,레포트,출석,중간,기말). 현재 {len(parts)}개.')
                continue

            name = parts[0]
            if not name:
                errors.append(f'라인 {i}: 이름이 비어있습니다.')
                continue

            field_labels = ['레포트', '출석', '중간시험', '기말시험']
            score_values = []
            row_ok = True
            for j, label in enumerate(field_labels):
                val, err = _parse_score(parts[j + 1], label)
                if err:
                    errors.append(f'라인 {i}: {err}')
                    row_ok = False
                    break
                score_values.append(val)

            if not row_ok:
                continue

            student = Student(
                course_id=course_id,
                name=name,
                report_score=score_values[0],
                attendance_score=score_values[1],
                midterm_score=score_values[2],
                final_score=score_values[3],
            )
            db.session.add(student)
            added += 1

        if added > 0:
            db.session.commit()
            recalculate_grades(course_id)
            flash(f'{added}명의 학생이 추가되었습니다.', 'success')
        else:
            db.session.rollback()

        for err in errors:
            flash(err, 'error')

        if added > 0:
            return redirect(url_for('course_detail', course_id=course_id))

    return render_template('bulk_add.html', course=course)


# ---------------------------------------------------------------------------
# Routes — Excel Import
# ---------------------------------------------------------------------------

@app.route('/course/<int:course_id>/import_excel', methods=['GET', 'POST'])
def import_excel(course_id):
    """Import students from .xlsx file (A=name, B=report, C=attendance, D=midterm, E=final)."""
    course = Course.query.get_or_404(course_id)

    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file or file.filename == '':
            flash('파일을 선택해주세요.', 'error')
            return render_template('import_excel.html', course=course)

        if not file.filename.lower().endswith('.xlsx'):
            flash('.xlsx 파일만 업로드할 수 있습니다.', 'error')
            return render_template('import_excel.html', course=course)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
        except Exception:
            app.logger.exception('Excel parse error')
            flash('파일을 읽는 중 오류가 발생했습니다. 올바른 .xlsx 파일인지 확인해주세요.', 'error')
            return render_template('import_excel.html', course=course)

        added = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row[:5]):
                continue

            name = str(row[0]).strip() if row[0] is not None else ''
            if not name:
                errors.append(f'행 {row_num}: 이름이 비어있습니다.')
                continue

            field_labels = ['레포트', '출석', '중간시험', '기말시험']
            score_values = []
            row_ok = True
            for j, label in enumerate(field_labels):
                raw = row[j + 1] if len(row) > j + 1 else None
                val, err = _parse_score(raw, label)
                if err:
                    errors.append(f'행 {row_num}: {err}')
                    row_ok = False
                    break
                score_values.append(val)

            if not row_ok:
                continue

            student = Student(
                course_id=course_id,
                name=name,
                report_score=score_values[0],
                attendance_score=score_values[1],
                midterm_score=score_values[2],
                final_score=score_values[3],
            )
            db.session.add(student)
            added += 1

        if added > 0:
            db.session.commit()
            recalculate_grades(course_id)
            flash(f'{added}명의 학생이 가져와졌습니다.', 'success')
        else:
            db.session.rollback()

        for err in errors:
            flash(err, 'error')

        if added > 0:
            return redirect(url_for('course_detail', course_id=course_id))

    return render_template('import_excel.html', course=course)


# ---------------------------------------------------------------------------
# Routes — Roster (명단 관리)
# ---------------------------------------------------------------------------

@app.route('/course/<int:course_id>/upload_roster', methods=['GET', 'POST'])
def upload_roster(course_id):
    """Upload a class roster Excel file (.xls/.xlsx) and store into CourseRoster."""
    course = Course.query.get_or_404(course_id)

    if request.method == 'POST':
        file = request.files.get('roster_file')
        if not file or file.filename == '':
            flash('파일을 선택해주세요.', 'error')
            return render_template('upload_roster.html', course=course)

        mode = request.form.get('mode', 'append')  # 'append' or 'replace'

        try:
            file_data = file.read()
            entries = parse_roster_file(file_data, file.filename)
        except ValueError as exc:
            flash(str(exc), 'error')
            return render_template('upload_roster.html', course=course)

        if not entries:
            flash('파일에 유효한 학생 데이터가 없습니다.', 'error')
            return render_template('upload_roster.html', course=course)

        if mode == 'replace':
            CourseRoster.query.filter_by(course_id=course_id).delete(synchronize_session='fetch')

        for entry in entries:
            roster_entry = CourseRoster(
                course_id=course_id,
                department=entry['department'] or None,
                name=entry['name'],
                student_number=entry['student_number'] or None,
            )
            db.session.add(roster_entry)

        db.session.commit()
        flash(f'{len(entries)}명 명단 등록 완료', 'success')
        return redirect(url_for('add_from_roster', course_id=course_id))

    return render_template('upload_roster.html', course=course)


@app.route('/course/<int:course_id>/clear_roster', methods=['POST'])
def clear_roster(course_id):
    """Delete all CourseRoster entries for the course."""
    course = Course.query.get_or_404(course_id)
    CourseRoster.query.filter_by(course_id=course_id).delete()
    db.session.commit()
    flash('명단이 초기화되었습니다.', 'success')
    return redirect(url_for('course_detail', course_id=course_id))


@app.route('/course/<int:course_id>/add_from_roster', methods=['GET', 'POST'])
def add_from_roster(course_id):
    """
    GET  — Show roster table with client-side filter; optionally show score entry
            form when ?roster_id=X is present.
    POST — Save a new Student from the score entry form.
    """
    course = Course.query.get_or_404(course_id)

    if request.method == 'POST':
        # Collect pre-filled identity fields
        name = request.form.get('name', '').strip()
        if not name:
            flash('학생 이름을 입력해주세요.', 'error')
            roster_id = request.form.get('roster_id', '')
            return redirect(url_for('add_from_roster', course_id=course_id, roster_id=roster_id))

        student_number = request.form.get('student_number', '').strip() or None
        department = request.form.get('department', '').strip() or None

        score_fields = [
            ('report_score',     '레포트'),
            ('attendance_score', '출석'),
            ('midterm_score',    '중간시험'),
            ('final_score',      '기말시험'),
        ]
        scores = {}
        for field, label in score_fields:
            val, err = _parse_score(request.form.get(field, ''), label)
            if err:
                flash(err, 'error')
                roster_id = request.form.get('roster_id', '')
                return redirect(url_for('add_from_roster', course_id=course_id, roster_id=roster_id))
            scores[field] = val

        student = Student(
            course_id=course_id,
            name=name,
            student_number=student_number,
            department=department,
            **scores,
        )
        db.session.add(student)
        db.session.commit()
        recalculate_grades(course_id)

        flash(f'{name} 학생이 추가되었습니다.', 'success')
        return redirect(url_for('course_detail', course_id=course_id))

    # GET — load roster, optionally pre-select one entry
    roster = CourseRoster.query.filter_by(course_id=course_id).order_by(
        CourseRoster.department.asc(), CourseRoster.name.asc()
    ).all()

    selected_roster = None
    roster_id = request.args.get('roster_id', '')
    if roster_id:
        try:
            selected_roster = db.session.get(CourseRoster, int(roster_id))
            # Only allow entries that belong to this course
            if selected_roster and selected_roster.course_id != course_id:
                selected_roster = None
        except (ValueError, TypeError):
            selected_roster = None

    # Build unique department list for the filter dropdown
    departments = sorted({r.department for r in roster if r.department})

    return render_template(
        'add_from_roster.html',
        course=course,
        roster=roster,
        selected_roster=selected_roster,
        departments=departments,
    )


# ---------------------------------------------------------------------------
# Routes — Statistics
# ---------------------------------------------------------------------------

@app.route('/course/<int:course_id>/statistics')
def statistics(course_id):
    """Per-course statistics: average, high, low, grade distribution."""
    course = Course.query.get_or_404(course_id)
    students = Student.query.filter_by(course_id=course_id).all()

    stats = {}
    if students:
        totals = [s.total_score for s in students if s.total_score is not None]
        if totals:
            stats['total_students'] = len(students)
            stats['average_score'] = round(sum(totals) / len(totals), 2)
            stats['highest_score'] = round(max(totals), 2)
            stats['lowest_score'] = round(min(totals), 2)

            grade_dist = {}
            for s in students:
                g = s.grade or 'F'
                grade_dist[g] = grade_dist.get(g, 0) + 1
            stats['grade_distribution'] = grade_dist

            stats['grade_percentages'] = {
                g: round(cnt / len(students) * 100, 1)
                for g, cnt in grade_dist.items()
            }

    grade_order = ['A+', 'A', 'B+', 'B', 'C+', 'C', 'D+', 'D', 'F']
    return render_template(
        'statistics.html',
        course=course,
        students=students,
        stats=stats,
        grade_order=grade_order,
    )


# ---------------------------------------------------------------------------
# App entry point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    with app.app_context():
        # NOTE (Step 2 upgrade): db.create_all() will NOT automatically add
        # student_number / department columns to an existing 'student' table,
        # nor will it create the new 'course_roster' table if the DB already
        # exists from Step 1.
        # If upgrading from Step 1, delete grade_management.db and restart.
        db.create_all()
    app.run(debug=True)
